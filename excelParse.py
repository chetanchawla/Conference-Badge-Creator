from openpyxl import *
from openpyxl.cell.cell import *

import sys

DATA_FILE = 'attendees.log'

def append_from_column(lst, ws, index, length):
    # for row in ws.values:
    #     for value in row:
    #         print(value)
    for i in range(1, length):
        lst.append(ws.cell(row=i, column=index).value)
        #print(lst) --debugging
    #print( ws.cell(row=i, column=).value )
    #[ lst.append( ws[index+str(i)].value ) for i in range(length) ]

def parse_data(config):
    with open(config) as f:
        import json
        data = json.loads(f.read())
    return data

def create_dictionary(names, surnames, occupations, universities, numOfRows):
    print ("[+] Writing everything to a list of dictionaries")
    full = []
    for i in range(numOfRows-1):
        #print(names) 
        #print(i)
        #for debugging
        #print("Name: ", names[i],"Surname: ", surnames[i],"Occupation: ", occupations[i], "University: ", universities[i])
        dictionary = {
            "Name": names[i],
            "Surname": surnames[i],
            "Occupation": occupations[i],
            "University": universities[i]
            }
        # print '[!] DEBUG'+str(names[i].decode("utf-8"))
        full.append(dictionary)
    return full

def parse_excelfile(ws, data_format, numOfRows):
    names = []
    surnames = []
    occupations = []
    universities = []

    # Loop through whole excel file
    print ("[+] Reading worksheet, appending Names/Surnames/Occupations")
    numOfColumns=ws.max_column + 1
    for column in range(1,numOfColumns):
        column_letter = get_column_letter(column)  # openpyxl function
        #print(column_letter) --for debugging
        if data_format[column_letter].lower() == 'name':
            append_from_column(names, ws, column, numOfRows)
        elif data_format[column_letter].lower() == 'surname':  
            append_from_column(surnames, ws, column, numOfRows)
        elif data_format[column_letter].lower() == 'occupation':  
            append_from_column(occupations, ws, column, numOfRows)
        elif data_format[column_letter].lower() == 'university':
            append_from_column(universities, ws, column, numOfRows)
        else:
            print "ERROR!"
            sys.exit(-1)

    return names, surnames, occupations, universities

def excel_to_dict(file_url, config_file):
    ws = load_workbook(filename=file_url).active
    numOfRows = ws.max_row + 1
    numOfColumns = ws.max_column + 1

    data_format = parse_data(config_file)
    names, surnames, occupations, universities = parse_excelfile(ws, data_format, numOfRows)
    dic = create_dictionary(names, surnames, occupations, universities, numOfRows)

    print ("[+] Writing dictionary to file.")
    with open(DATA_FILE, 'w') as f:
        f.write(str(dic))
    print("[!] Done!")

